# backend/admin/sold_routes.py
import io
import os
from datetime import datetime, timedelta
from decimal import Decimal

from flask import (
    Blueprint, request, render_template, send_file, url_for, redirect, flash, current_app
)
from flask_login import login_required
from flask_mail import Message

from . import admin_bp  # zachovĂˇno
from backend.extensions import db, mail
from backend.models import SoldProduct
from backend.invoicing import build_invoice_pdf_bytes  # vyuĹľijeme existujĂ­cĂ­ generĂˇtor FA


# -----------------------------
# PomocnĂ© funkce â€“ parsovĂˇnĂ­/ÄŤĂ­sla
# -----------------------------

def _parse_date(s: str):
    """BezpeÄŤnĂ© parsovĂˇnĂ­ data ve formĂˇtu YYYY-MM-DD, vracĂ­ datetime nebo None."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None


def _to_float(v) -> float:
    """RobustnĂ­ pĹ™evod na float (podpora Decimal, Ĺ™etÄ›zcĹŻ s ÄŤĂˇrkou atd.)."""
    if v is None:
        return 0.0
    if isinstance(v, float):
        return v
    if isinstance(v, int):
        return float(v)
    if isinstance(v, Decimal):
        return float(v)
    try:
        s = str(v).strip().replace(",", ".")
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _get_first_attr(obj, *names):
    """VrĂˇtĂ­ prvnĂ­ nenull/neen-prĂˇzdnĂ˝ atribut z dodanĂ˝ch nĂˇzvĹŻ, jinak None."""
    for n in names:
        if hasattr(obj, n):
            val = getattr(obj, n)
            if val not in (None, ""):
                return val
    return None


def _guess_unit_price_czk(sp) -> float:
    """Najdi nejpravdÄ›podobnÄ›jĹˇĂ­ pole s jednotkovou cenou (CZK)."""
    cand = _get_first_attr(
        sp,
        "unit_price_czk", "price_czk",
        "unit_price", "price",
        "unit", "unit_czk", "sold_unit_price_czk", "final_price_czk",
    )
    return round(_to_float(cand), 2)


def _guess_quantity(sp) -> float:
    """Najdi mnoĹľstvĂ­; fallback na 1.0, pokud nic rozumnĂ©ho nenalezeno."""
    cand = _get_first_attr(sp, "quantity", "qty", "count", "pieces", "amount_units")
    q = _to_float(cand)
    return q if q > 0 else 1.0


def _guess_total_czk(sp) -> float:
    """
    Najdi celkovou cenu; kdyĹľ chybĂ­ nebo je 0/None, dopoÄŤĂ­tej unit * qty.
    """
    cand = _get_first_attr(sp, "total_czk", "total_price_czk", "amount_czk", "amount")
    total = _to_float(cand)
    if total > 0:
        return round(total, 2)
    unit = _guess_unit_price_czk(sp)
    qty = _guess_quantity(sp)
    return round(unit * qty, 2)


def _sold_datetime(sp):
    """VraĹĄ datetime pro setĹ™Ă­dÄ›nĂ­/filtr (preferuj sold_at, jinak created_at)."""
    return _get_first_attr(sp, "sold_at", "created_at")


def _base_query(filters: dict):
    """
    ZĂˇkladnĂ­ dotaz pro SoldProduct s podporou ?from & ?to (vÄŤetnÄ›).
    TĹ™Ă­dĂ­ dle sold_at desc, pokud je k dispozici; jinak dle id desc.
    """
    q = SoldProduct.query
    d_from = _parse_date(filters.get("from"))
    d_to = _parse_date(filters.get("to"))

    if hasattr(SoldProduct, "sold_at"):
        if d_from:
            q = q.filter(SoldProduct.sold_at >= d_from)
        if d_to:
            q = q.filter(SoldProduct.sold_at < (d_to + timedelta(days=1)))
        q = q.order_by(SoldProduct.sold_at.desc())
    else:
        if hasattr(SoldProduct, "id"):
            q = q.order_by(SoldProduct.id.desc())

    return q


def _to_row_dict(sp: SoldProduct):
    """MapovĂˇnĂ­ zĂˇznamu na Ĺ™Ăˇdek do Ĺˇablony/exportĹŻ (vÄŤetnÄ› dopoÄŤtu cen)."""
    sold_dt = _sold_datetime(sp)
    return {
        "id": getattr(sp, "id", None),
        "order_id": _get_first_attr(sp, "order_id", "order_code"),
        "product_name": _get_first_attr(sp, "product_name", "name"),
        "quantity": _guess_quantity(sp),
        "unit_price_czk": _guess_unit_price_czk(sp),
        "total_czk": _guess_total_czk(sp),
        "status": getattr(sp, "status", None),
        "customer_email": _get_first_attr(sp, "customer_email", "email"),
        "vs": getattr(sp, "vs", None),
        "sold_at": sold_dt,
    }


# -----------------------------
# Seznam + filtry
# -----------------------------

@admin_bp.route("/sold")
# # # # @login_required  # dočasně vypnuto (dočasně vypnuto)
def sold_products():
    """
    Seznam prodanĂ˝ch s filtry ?from=YYYY-MM-DD&to=YYYY-MM-DD
    a souhrnem (poÄŤet, suma).
    """
    filters = {
        "from": request.args.get("from", "").strip(),
        "to": request.args.get("to", "").strip(),
    }

    q = _base_query(filters)
    items = q.all()

    # ManuĂˇlnĂ­ filtr dle datumu, pokud model nemĂˇ sold_at sloupec
    d_from = _parse_date(filters.get("from"))
    d_to_excl = _parse_date(filters.get("to"))
    if d_to_excl:
        d_to_excl = d_to_excl + timedelta(days=1)

    if not hasattr(SoldProduct, "sold_at") and (d_from or d_to_excl):
        filtered = []
        for sp in items:
            dt = _sold_datetime(sp)
            if isinstance(dt, datetime):
                if d_from and dt < d_from:
                    continue
                if d_to_excl and dt >= d_to_excl:
                    continue
            filtered.append(sp)
        items = filtered

    rows = [_to_row_dict(sp) for sp in items]

    # Souhrn
    count = len(rows)
    total_amount = round(sum(_to_float(r["total_czk"]) for r in rows), 2)

    # ĹazenĂ­ (nejnovÄ›jĹˇĂ­ nahoĹ™e)
    rows.sort(key=lambda r: r["sold_at"] or datetime.min, reverse=True)

    return render_template(
        "admin/sold/list.html",
        filters=filters,
        rows=rows,
        summary={"count": count, "total_amount": total_amount},
    )


# -----------------------------
# Export: Excel (XLSX)
# -----------------------------

@admin_bp.route("/sold/export.xlsx")
# # # # @login_required  # dočasně vypnuto (dočasně vypnuto)
def sold_export_xlsx():
    """
    Export do Excelu s aktuĂˇlnĂ­mi filtry.
    VyĹľaduje knihovnu openpyxl (pip install openpyxl).
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("ChybĂ­ balĂ­k openpyxl. Nainstaluj: pip install openpyxl", "danger")
        return redirect(url_for("admin.sold_products", **request.args))

    filters = {
        "from": request.args.get("from", "").strip(),
        "to": request.args.get("to", "").strip(),
    }
    q = _base_query(filters)
    items = q.all()

    # ManuĂˇlnĂ­ datum filtr, pokud nenĂ­ sold_at
    d_from = _parse_date(filters.get("from"))
    d_to_excl = _parse_date(filters.get("to"))
    if d_to_excl:
        d_to_excl = d_to_excl + timedelta(days=1)
    if not hasattr(SoldProduct, "sold_at") and (d_from or d_to_excl):
        filtered = []
        for sp in items:
            dt = _sold_datetime(sp)
            if isinstance(dt, datetime):
                if d_from and dt < d_from:
                    continue
                if d_to_excl and dt >= d_to_excl:
                    continue
            filtered.append(sp)
        items = filtered

    data = [_to_row_dict(sp) for sp in items]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ProdanĂ© poloĹľky"

    headers = [
        "ID", "ObjednĂˇvka", "Produkt", "MnoĹľstvĂ­",
        "Cena/ks (KÄŤ)", "Celkem (KÄŤ)", "Status",
        "E-mail", "VS", "ProdanĂ©"
    ]
    ws.append(headers)

    for r in data:
        ws.append([
            r["id"], r["order_id"], r["product_name"], r["quantity"],
            r["unit_price_czk"], r["total_czk"], r["status"],
            r["customer_email"], r["vs"],
            r["sold_at"].strftime("%Y-%m-%d %H:%M") if isinstance(r["sold_at"], datetime) else ""
        ])

    # Auto ĹˇĂ­Ĺ™ky
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"sold_{filters.get('from','')}_{filters.get('to','')}.xlsx".replace("__", "_").strip("_")
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename or "sold.xlsx",
    )


# -----------------------------
# PDF â€“ pomocnĂˇ registrace CZ fontu
# -----------------------------

def _register_cz_fonts(pdfmetrics, TTFont):
    """
    ZkusĂ­ zaregistrovat Unicode TTF font pro diakritiku.
    Preferuje DejaVu Sans, pak Noto Sans, nakonec Windows Arial.
    VracĂ­ (regular_name, bold_name) nebo (None, None), kdyĹľ nic nenaĹˇel.
    """
    root = current_app.root_path  # backend/
    candidates = [
        # DejaVu Sans
        (os.path.join(root, "static", "fonts", "DejaVuSans.ttf"),
         os.path.join(root, "static", "fonts", "DejaVuSans-Bold.ttf"),
         "DejaVuSans", "DejaVuSans-Bold"),
        # Noto Sans
        (os.path.join(root, "static", "fonts", "NotoSans-Regular.ttf"),
         os.path.join(root, "static", "fonts", "NotoSans-Bold.ttf"),
         "NotoSans", "NotoSans-Bold"),
        # Windows Arial (poslednĂ­ zĂˇchrana)
        (r"C:\Windows\Fonts\arial.ttf",
         r"C:\Windows\Fonts\arialbd.ttf",
         "ArialTT", "ArialTT-Bold"),
    ]

    for reg_path, bold_path, reg_name, bold_name in candidates:
        if os.path.isfile(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(reg_name, reg_path))
                if os.path.isfile(bold_path):
                    pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                else:
                    bold_name = reg_name  # fallback na regular, pokud bold chybĂ­
                return reg_name, bold_name
            except Exception:
                continue
    return None, None


# -----------------------------
# Export: PDF report seznamu (s CZ fontem)
# -----------------------------

@admin_bp.route("/sold/export.pdf")
# # # # @login_required  # dočasně vypnuto (dočasně vypnuto)
def sold_export_pdf():
    """
    JednoduchĂ˝ PDF report (seznam). VyĹľaduje reportlab (pip install reportlab).
    Registruje Unicode TTF font (DejaVu/Noto/Arial), aby fungovala ÄŤeĹˇtina.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        flash("ChybĂ­ balĂ­k reportlab. Nainstaluj: pip install reportlab", "danger")
        return redirect(url_for("admin.sold_products", **request.args))

    # Registrace CZ fontu
    regular_font, bold_font = _register_cz_fonts(pdfmetrics, TTFont)
    if not regular_font:
        flash(
            "Pro korektnĂ­ ÄŤeĹˇtinu v PDF vloĹľ TTF font do backend/static/fonts/ "
            "(napĹ™. DejaVuSans.ttf a DejaVuSans-Bold.ttf). ZatĂ­m pouĹľĂ­vĂˇm Helvetica.",
            "warning",
        )
    title_font = bold_font or "Helvetica-Bold"
    text_font = regular_font or "Helvetica"

    # Data
    filters = {
        "from": request.args.get("from", "").strip(),
        "to": request.args.get("to", "").strip(),
    }
    q = _base_query(filters)
    items = q.all()

    # ManuĂˇlnĂ­ datum filtr, pokud nenĂ­ sold_at
    d_from = _parse_date(filters.get("from"))
    d_to_excl = _parse_date(filters.get("to"))
    if d_to_excl:
        d_to_excl = d_to_excl + timedelta(days=1)
    if not hasattr(SoldProduct, "sold_at") and (d_from or d_to_excl):
        filtered = []
        for sp in items:
            dt = _sold_datetime(sp)
            if isinstance(dt, datetime):
                if d_from and dt < d_from:
                    continue
                if d_to_excl and dt >= d_to_excl:
                    continue
            filtered.append(sp)
        items = filtered

    data = [_to_row_dict(sp) for sp in items]

    # PDF
    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    # Header
    c.setFont(title_font, 14)
    c.drawString(20 * mm, (h - 20 * mm), "Report â€“ ProdanĂ© poloĹľky")

    c.setFont(text_font, 10)
    c.drawString(20 * mm, (h - 27 * mm), f"Filtr: od {filters.get('from') or '-'} do {filters.get('to') or '-'}")
    c.drawString(20 * mm, (h - 33 * mm), f"PoÄŤet: {len(data)}")

    # ZĂˇhlavĂ­ tabulky
    def draw_header(y_pos):
        c.setFont(title_font, 9)
        c.drawString(20 * mm, y_pos, "ID")
        c.drawString(35 * mm, y_pos, "ObjednĂˇvka")
        c.drawString(65 * mm, y_pos, "Produkt")
        c.drawString(120 * mm, y_pos, "Celkem KÄŤ")
        c.drawString(150 * mm, y_pos, "ProdanĂ©")
        return y_pos - 6 * mm

    y = h - 45 * mm
    y = draw_header(y)

    # ĹĂˇdky
    c.setFont(text_font, 9)
    for r in data:
        if y < 20 * mm:
            c.showPage()
            c.setFont(text_font, 9)
            y = h - 20 * mm
            y = draw_header(y)

        c.drawString(20 * mm, y, str(r["id"] or ""))
        c.drawString(35 * mm, y, str(r["order_id"] or ""))
        c.drawString(65 * mm, y, str((r["product_name"] or "")[:32]))
        c.drawRightString(145 * mm, y, f"{_to_float(r['total_czk']):.2f}")
        dt = r["sold_at"]
        c.drawString(150 * mm, y, dt.strftime("%Y-%m-%d") if isinstance(dt, datetime) else "")
        y -= 6 * mm

    c.showPage()
    c.save()
    bio.seek(0)

    filename = f"sold_{filters.get('from','')}_{filters.get('to','')}.pdf".replace("__", "_").strip("_")
    return send_file(
        bio,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename or "sold.pdf",
    )


# -----------------------------
# Faktura â€“ nĂˇhled (PDF)
# -----------------------------

@admin_bp.route("/invoice/<int:sold_id>.pdf")
# # # # @login_required  # dočasně vypnuto (dočasně vypnuto)
def invoice_preview_pdf(sold_id: int):
    """
    NĂˇhled faktury pro konkrĂ©tnĂ­ SoldProduct (PDF).
    VyuĹľĂ­vĂˇ build_invoice_pdf_bytes(sp) z backend.invoicing.
    """
    sp = SoldProduct.query.get_or_404(sold_id)
    pdf_bytes = build_invoice_pdf_bytes(sp)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"invoice_{sold_id}.pdf",
    )


# -----------------------------
# OdeslĂˇnĂ­ faktury e-mailem (jedna poloĹľka)
# -----------------------------

@admin_bp.route("/invoice/<int:sold_id>/email", methods=["POST"])
# # # # @login_required  # dočasně vypnuto (dočasně vypnuto)
def invoice_send_email(sold_id: int):
    """
    PoĹˇle fakturu e-mailem na 'customer_email' (pokud existuje).
    PĹ™iklĂˇdĂˇ PDF generovanĂ© pĹ™es build_invoice_pdf_bytes.
    """
    sp = SoldProduct.query.get_or_404(sold_id)
    to_addr = getattr(sp, "customer_email", None)

    if not to_addr:
        flash("U zĂˇznamu nenĂ­ vyplnÄ›n e-mail zĂˇkaznĂ­ka.", "warning")
        return redirect(url_for("admin.sold_products"))

    pdf_bytes = build_invoice_pdf_bytes(sp)

    try:
        subject = f"Faktura #{getattr(sp, 'id', '')} â€“ NĂˇramkovĂˇ MĂłda"
        body = (
            "DobrĂ˝ den,\n\n"
            "v pĹ™Ă­loze zasĂ­lĂˇme fakturu k VaĹˇĂ­ objednĂˇvce.\n"
            "DÄ›kujeme za nĂˇkup.\n\n"
            "S pozdravem\nNĂˇramkovĂˇ MĂłda"
        )
        msg = Message(subject=subject, recipients=[to_addr], body=body)
        msg.attach(
            filename=f"invoice_{sp.id}.pdf",
            content_type="application/pdf",
            data=pdf_bytes,
        )
        mail.send(msg)
        flash(f"E-mail s fakturou byl odeslĂˇn na {to_addr}.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"NepodaĹ™ilo se odeslat e-mail: {e}", "danger")

    return redirect(url_for("admin.sold_products"))


# =============================
# NOVĂ‰: OdeslĂˇnĂ­ faktury podle ORDER_ID (pro â€žauto po zaplacenĂ­â€ś)
#  - PouĹľij z adminu/servisu po zmÄ›nÄ› stavu na 'paid' / 'zaplaceno'
#  - VracĂ­ dict s vĂ˝sledkem; volĂˇnĂ­ je idempotentnĂ­ v tom smyslu,
#    Ĺľe kdyĹľ pro objednĂˇvku nenĂ­ ĹľĂˇdnĂ˝ SoldProduct, jen vrĂˇtĂ­ info.
# =============================

def send_invoice_for_order(order_id: int) -> dict:
    """
    Najde prvnĂ­ SoldProduct pro danĂ˝ order_id, vygeneruje PDF a:
      - pokud uĹľ je invoice oznaÄŤenĂˇ jako odeslanĂˇ (invoice_sent_at), NIC neposĂ­lĂˇ (idempotentnĂ­),
      - jinak poĹˇle e-mail, uloĹľĂ­ PDF na disk a zapĂ­Ĺˇe info do DB (pokud mĂˇ model pole).
    """
    sp = (SoldProduct.query
          .filter(SoldProduct.order_id == order_id)
          .order_by(SoldProduct.id.asc())
          .first())

    if not sp:
        return {"ok": False, "error": "No SoldProduct for this order_id."}

    # 1) idempotence: pokud existuje flag, nepĹ™eposĂ­lat
    already_sent = False
    if hasattr(sp, "invoice_sent_at") and getattr(sp, "invoice_sent_at", None):
        already_sent = True

    # 2) vygeneruj PDF
    pdf_bytes = build_invoice_pdf_bytes(sp)

    # 3) uloĹľ PDF na disk (aĹĄ mĂˇĹˇ archiv)
    try:
        inv_dir = os.path.join(current_app.root_path, "static", "invoices")
        os.makedirs(inv_dir, exist_ok=True)
        filename = f"invoice_order_{getattr(sp, 'order_id', 'x')}_sold_{sp.id}.pdf"
        file_path = os.path.join(inv_dir, filename)
        with open(file_path, "wb") as f:
            f.write(pdf_bytes)
        saved_rel = f"/static/invoices/{filename}"
    except Exception:
        saved_rel = None

    result = {
        "ok": True,
        "emailed": False,
        "sold_id": getattr(sp, "id", None),
        "saved_path": saved_rel,
        "already_sent": already_sent,
    }

    # 4) pokud uĹľ bylo dĹ™Ă­ve odeslĂˇno, skonÄŤi (jen informativnÄ› vrĂˇtĂ­me)
    if already_sent:
        return result

    # 5) poĹˇli e-mail (jen pokud mĂˇme adresu)
    to_addr = getattr(sp, "customer_email", None) or getattr(sp, "email", None)
    if not to_addr:
        return result

    try:
        subject = f"Faktura #{getattr(sp, 'id', '')} â€“ NĂˇramkovĂˇ MĂłda"
        body = (
            "DobrĂ˝ den,\n\n"
            "v pĹ™Ă­loze zasĂ­lĂˇme fakturu k VaĹˇĂ­ objednĂˇvce.\n"
            "DÄ›kujeme za nĂˇkup.\n\n"
            "S pozdravem\nNĂˇramkovĂˇ MĂłda"
        )
        msg = Message(subject=subject, recipients=[to_addr], body=body)
        # pĹ™iloĹľĂ­me stejnĂ© PDF, kterĂ© jsme uloĹľili
        msg.attach(
            filename=os.path.basename(saved_rel) if saved_rel else f"invoice_{sp.id}.pdf",
            content_type="application/pdf",
            data=pdf_bytes,
        )
        from backend.extensions import mail  # lokĂˇlnĂ­ import, aby se nerozbilo jinde
        mail.send(msg)
        result["emailed"] = True

        # 6) zapĂ­Ĺˇeme do DB flag (pokud model mĂˇ sloupce)
        dirty = False
        if hasattr(sp, "invoice_sent_at") and not getattr(sp, "invoice_sent_at", None):
            sp.invoice_sent_at = datetime.utcnow()
            dirty = True
        if saved_rel and hasattr(sp, "invoice_filename") and not getattr(sp, "invoice_filename", None):
            sp.invoice_filename = saved_rel
            dirty = True
        if dirty:
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
    except Exception as e:
        current_app.logger.exception(f"[invoice] send failed for order_id={order_id}: {e}")
        return {"ok": False, "error": str(e), **result}

    return result


# -----------------------------
# VOLITELNĂ‰: ruÄŤnĂ­ spuĹˇtÄ›nĂ­ podle ORDER_ID z UI
# -----------------------------

@admin_bp.route("/sold/order/<int:order_id>/email", methods=["POST"])
# # # # @login_required  # dočasně vypnuto (dočasně vypnuto)
def invoice_send_email_for_order(order_id: int):
    """
    RuÄŤnĂ­ odeslĂˇnĂ­ faktury podle order_id (vezme prvnĂ­ SoldProduct tĂ© objednĂˇvky).
    HodĂ­ se pro rychlĂ© â€žpĹ™eposlĂˇnĂ­â€ś, kdyĹľ doĹˇla zmÄ›na e-mailu apod.
    """
    res = send_invoice_for_order(order_id)
    if res.get("ok") and res.get("emailed"):
        flash(f"Faktura byla odeslĂˇna na {res.get('to')}.", "success")
    elif res.get("ok") and not res.get("emailed"):
        flash("Faktura nebyla odeslĂˇna â€“ u objednĂˇvky nenĂ­ e-mail zĂˇkaznĂ­ka.", "warning")
    else:
        flash(f"Fakturu se nepodaĹ™ilo odeslat: {res.get('error', 'neznĂˇmĂˇ chyba')}", "danger")

    # zpÄ›t kamkoliv â€“ ideĂˇlnÄ› na detail objednĂˇvky, pokud mĂˇĹˇ
    return redirect(request.referrer or url_for("admin.sold_products"))






